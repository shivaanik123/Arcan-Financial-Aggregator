import streamlit as st
from PyPDF2 import PdfMerger
import pdfplumber
import io
import json
import re
from pathlib import Path
from collections import defaultdict
from openpyxl import Workbook, load_workbook
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors

st.set_page_config(page_title="Arcan Financial Report Aggregator", layout="centered")

# Box configuration (use environment variables in production)
import os
import psycopg2
BOX_CLIENT_ID = os.environ.get("BOX_CLIENT_ID", "bfw6aqc5eaezh292nh638mss04hzhpxm")
BOX_CLIENT_SECRET = os.environ.get("BOX_CLIENT_SECRET", "J9ao1WBpsjbU4QUBPSTkq1vMxeNgHtGf")
BOX_ROOT_FOLDER_ID = os.environ.get("BOX_ROOT_FOLDER_ID", "7627162890")
BOX_REDIRECT_URI = os.environ.get("BOX_REDIRECT_URI", "http://localhost:8501")
DATABASE_URL = os.environ.get("DATABASE_URL", "postgresql://postgres:urMzbfSCtHlGoJWoNnqSYALFYImWQplu@postgres.railway.internal:5432/railway")

# Initialize database
def init_db():
    try:
        conn = psycopg2.connect(DATABASE_URL)
        cur = conn.cursor()
        # Check if old schema exists (username column) and migrate if needed
        cur.execute("""
            SELECT column_name FROM information_schema.columns
            WHERE table_name = 'box_tokens' AND column_name = 'username'
        """)
        if cur.fetchone():
            # Old schema exists, drop and recreate
            cur.execute("DROP TABLE box_tokens")

        cur.execute("""
            CREATE TABLE IF NOT EXISTS box_tokens (
                box_user_id VARCHAR(100) PRIMARY KEY,
                box_user_name VARCHAR(255),
                box_user_email VARCHAR(255),
                access_token TEXT,
                refresh_token TEXT
            )
        """)
        conn.commit()
        cur.close()
        conn.close()
    except Exception as e:
        st.error(f"Database error: {e}")

init_db()

MONTH_NAMES = {
    "01": "January", "02": "February", "03": "March", "04": "April",
    "05": "May", "06": "June", "07": "July", "08": "August",
    "09": "September", "10": "October", "11": "November", "12": "December"
}

def is_t12_or_ytd(file_bytes):
    """Determine if a 12-month statement is T-12 or YTD based on date range.

    T-12: spans 12 months (trailing 12)
    YTD: starts in January, less than 12 months

    Returns 'T-12' or 'YTD' or None if can't determine.
    """
    try:
        with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
            if len(pdf.pages) > 0:
                text = pdf.pages[0].extract_text() or ""

                # Look for date range patterns like "Jan 2026 - Dec 2026" or "February 2025 - January 2026"
                month_pattern = r'(Jan(?:uary)?|Feb(?:ruary)?|Mar(?:ch)?|Apr(?:il)?|May|Jun(?:e)?|Jul(?:y)?|Aug(?:ust)?|Sep(?:tember)?|Oct(?:ober)?|Nov(?:ember)?|Dec(?:ember)?)'
                date_range = re.search(rf'{month_pattern}\s*\d{{4}}\s*[-–—to]+\s*{month_pattern}\s*\d{{4}}', text, re.IGNORECASE)

                if date_range:
                    start_month = date_range.group(1).lower()[:3]
                    # If starts in January, it's likely YTD
                    if start_month == 'jan':
                        return 'YTD'
                    else:
                        return 'T-12'

    except Exception:
        pass

    return None

def extract_property_info(file_bytes):
    """Extract property name and code from PDF content.

    Looks for pattern like 'The Turn (turn)' and extracts both 'The Turn' and 'turn'.
    Returns (property_name, property_code) tuple.
    """
    try:
        with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
            if len(pdf.pages) > 0:
                # Get text from first page
                text = pdf.pages[0].extract_text() or ""

                # Look for pattern: "Property Name (code)" at start of lines
                # This pattern appears in report headers
                match = re.search(r'^([A-Za-z][A-Za-z0-9\s&\'-]+?)\s*\(([a-z0-9_]+)\)', text, re.MULTILINE)
                if match:
                    property_name = match.group(1).strip().title()
                    property_code = match.group(2).strip().lower()
                    return (property_name, property_code)

                # Fallback: look for property name in common header patterns
                # Check for "Property = X" or similar patterns
                match = re.search(r'Property\s*[=:]\s*([A-Za-z][A-Za-z0-9\s&\'-]+?)(?:\s*\(|\s*$|\s*Page)', text)
                if match:
                    return (match.group(1).strip().title(), None)

    except Exception as e:
        st.warning(f"Could not extract property info: {e}")

    return (None, None)

def extract_content_from_pdf(file_bytes):
    """Extract content from a PDF file - tries tables first, then text."""
    all_rows = []
    try:
        with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
            for page_num, page in enumerate(pdf.pages):
                # Try to extract tables first
                tables = page.extract_tables()
                if tables:
                    for table in tables:
                        if table:
                            for row in table:
                                if row and any(cell for cell in row):  # Skip empty rows
                                    all_rows.append(row)
                            all_rows.append([])  # Empty row between tables
                else:
                    # No tables found - extract text line by line
                    text = page.extract_text()
                    if text:
                        lines = text.split('\n')
                        for line in lines:
                            if line.strip():
                                # Split by multiple spaces to create columns
                                parts = re.split(r'\s{2,}', line.strip())
                                all_rows.append(parts)
                        all_rows.append([])  # Empty row between pages
    except Exception as e:
        st.warning(f"Could not extract content: {e}")
    return all_rows

def excel_to_pdf(excel_bytes):
    """Convert Excel file to PDF using LibreOffice (preserves formatting)."""
    import subprocess
    import tempfile
    import os

    # LibreOffice paths to try
    libreoffice_paths = [
        '/Applications/LibreOffice.app/Contents/MacOS/soffice',
        '/usr/local/bin/soffice',
        '/usr/bin/soffice',
        'soffice',
        'libreoffice'
    ]

    # Find LibreOffice
    soffice_path = None
    for path in libreoffice_paths:
        try:
            result = subprocess.run([path, '--version'], capture_output=True, timeout=5)
            if result.returncode == 0:
                soffice_path = path
                break
        except:
            continue

    if not soffice_path:
        raise Exception("LibreOffice not found. Install with: brew install --cask libreoffice")

    # Create temp directory for conversion
    with tempfile.TemporaryDirectory() as temp_dir:
        # Write Excel to temp file
        excel_path = os.path.join(temp_dir, "input.xlsx")
        with open(excel_path, 'wb') as f:
            f.write(excel_bytes)

        # Convert to PDF using LibreOffice
        result = subprocess.run([
            soffice_path,
            '--headless',
            '--convert-to', 'pdf',
            '--outdir', temp_dir,
            excel_path
        ], capture_output=True, timeout=60)

        if result.returncode != 0:
            raise Exception(f"LibreOffice conversion failed: {result.stderr.decode()}")

        # Read the PDF output
        pdf_path = os.path.join(temp_dir, "input.pdf")
        if not os.path.exists(pdf_path):
            raise Exception("PDF output not created")

        with open(pdf_path, 'rb') as f:
            return f.read()


def merge_excel_files(t12_bytes, ytd_bytes, gl_bytes):
    """Merge three Excel files into one workbook with 3 sheets, preserving exact format."""
    from copy import copy
    from openpyxl.cell.cell import MergedCell
    from openpyxl.utils import get_column_letter
    import zipfile

    # Start with T-12 as base - load it directly to preserve exact formatting
    output = io.BytesIO()

    with zipfile.ZipFile(io.BytesIO(t12_bytes), 'r') as t12_zip:
        with zipfile.ZipFile(output, 'w', zipfile.ZIP_DEFLATED) as out_zip:
            for item in t12_zip.namelist():
                data = t12_zip.read(item)
                if item == 'xl/workbook.xml':
                    data = data.replace(b'Sheet1', b'T-12 Statement')
                    data = data.replace(b'Sheet 1', b'T-12 Statement')
                out_zip.writestr(item, data)

    output.seek(0)
    wb = load_workbook(output)
    wb.active.title = "T-12 Statement"

    def add_sheet_full_copy(source_bytes, sheet_name):
        """Add sheet with comprehensive formatting copy."""
        try:
            source_wb = load_workbook(io.BytesIO(source_bytes))
            source_ws = source_wb.active
            ws = wb.create_sheet(sheet_name)

            # Copy sheet properties
            ws.sheet_format = copy(source_ws.sheet_format)
            ws.sheet_properties = copy(source_ws.sheet_properties)

            # Copy all cells BEFORE merging
            for row in source_ws.iter_rows():
                for cell in row:
                    if isinstance(cell, MergedCell):
                        continue
                    new_cell = ws.cell(row=cell.row, column=cell.column)
                    new_cell.value = cell.value
                    if cell.has_style:
                        new_cell.font = copy(cell.font)
                        new_cell.border = copy(cell.border)
                        new_cell.fill = copy(cell.fill)
                        new_cell.number_format = cell.number_format
                        new_cell.protection = copy(cell.protection)
                        new_cell.alignment = copy(cell.alignment)

            # Apply merged cells
            for merged_range in list(source_ws.merged_cells.ranges):
                ws.merge_cells(str(merged_range))

            # Copy column dimensions
            for key, dim in source_ws.column_dimensions.items():
                ws.column_dimensions[key] = copy(dim)

            # Copy row dimensions
            for key, dim in source_ws.row_dimensions.items():
                ws.row_dimensions[key] = copy(dim)

            # Copy freeze panes
            ws.freeze_panes = source_ws.freeze_panes

            # Copy page setup
            ws.page_margins = copy(source_ws.page_margins)
            ws.page_setup = copy(source_ws.page_setup)

            source_wb.close()
        except Exception as e:
            if sheet_name not in wb.sheetnames:
                ws = wb.create_sheet(sheet_name)
            ws.cell(row=1, column=1, value=f"Error: {str(e)}")

    if ytd_bytes:
        add_sheet_full_copy(ytd_bytes, "YTD Statement")
    if gl_bytes:
        add_sheet_full_copy(gl_bytes, "General Ledger")

    # Save final output
    final_output = io.BytesIO()
    wb.save(final_output)
    final_output.seek(0)
    return final_output.getvalue()

def get_box_user_info(access_token):
    """Get the current Box user's info."""
    import requests
    response = requests.get(
        "https://api.box.com/2.0/users/me",
        headers={"Authorization": f"Bearer {access_token}"}
    )
    if response.status_code == 200:
        data = response.json()
        return {
            "id": data.get("id"),
            "name": data.get("name"),
            "email": data.get("login")
        }
    return None

def save_tokens(access_token, refresh_token, box_user_id=None, box_user_name=None, box_user_email=None):
    """Save tokens to database for the Box user."""
    if not box_user_id:
        box_user_id = st.session_state.get("box_user_id")
    if not box_user_id:
        return
    try:
        conn = psycopg2.connect(DATABASE_URL)
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO box_tokens (box_user_id, box_user_name, box_user_email, access_token, refresh_token)
            VALUES (%s, %s, %s, %s, %s)
            ON CONFLICT (box_user_id) DO UPDATE SET
                box_user_name = EXCLUDED.box_user_name,
                box_user_email = EXCLUDED.box_user_email,
                access_token = EXCLUDED.access_token,
                refresh_token = EXCLUDED.refresh_token
        """, (box_user_id, box_user_name, box_user_email, access_token, refresh_token))
        conn.commit()
        cur.close()
        conn.close()
    except Exception as e:
        st.error(f"Error saving tokens: {e}")

def load_tokens_by_id(box_user_id):
    """Load tokens from database for a specific Box user."""
    if not box_user_id:
        return None
    try:
        conn = psycopg2.connect(DATABASE_URL)
        cur = conn.cursor()
        cur.execute("SELECT access_token, refresh_token, box_user_name, box_user_email FROM box_tokens WHERE box_user_id = %s", (box_user_id,))
        row = cur.fetchone()
        cur.close()
        conn.close()
        if row:
            return {"access_token": row[0], "refresh_token": row[1], "name": row[2], "email": row[3]}
    except Exception as e:
        st.error(f"Error loading tokens: {e}")
    return None

def load_tokens():
    """Load tokens from database for the current session user."""
    box_user_id = st.session_state.get("box_user_id")
    return load_tokens_by_id(box_user_id)

def delete_tokens():
    """Delete tokens for the current Box user."""
    box_user_id = st.session_state.get("box_user_id")
    if not box_user_id:
        return
    try:
        conn = psycopg2.connect(DATABASE_URL)
        cur = conn.cursor()
        cur.execute("DELETE FROM box_tokens WHERE box_user_id = %s", (box_user_id,))
        conn.commit()
        cur.close()
        conn.close()
    except Exception as e:
        st.error(f"Error deleting tokens: {e}")

def get_box_client(access_token):
    """Get Box client with access token."""
    from box_sdk_gen import BoxClient, BoxOAuth, OAuthConfig, AccessToken

    config = OAuthConfig(client_id=BOX_CLIENT_ID, client_secret=BOX_CLIENT_SECRET)
    auth = BoxOAuth(config=config)
    token = AccessToken(access_token=access_token)
    client = BoxClient(auth=auth).with_custom_base_urls(base_url="https://api.box.com")

    # Use the access token directly
    from box_sdk_gen import BoxCCGAuth, CCGConfig
    from box_sdk_gen.networking.auth import Authentication

    class TokenAuth(Authentication):
        def __init__(self, access_token):
            self._access_token = access_token

        def retrieve_token(self, network_session=None):
            return AccessToken(access_token=self._access_token)

    return BoxClient(auth=TokenAuth(access_token))

def exchange_code_for_tokens(code):
    """Exchange authorization code for tokens."""
    import requests

    response = requests.post(
        "https://api.box.com/oauth2/token",
        data={
            "grant_type": "authorization_code",
            "code": code,
            "client_id": BOX_CLIENT_ID,
            "client_secret": BOX_CLIENT_SECRET,
            "redirect_uri": BOX_REDIRECT_URI,
        }
    )

    if response.status_code == 200:
        data = response.json()
        return data["access_token"], data["refresh_token"]
    else:
        raise Exception(f"Failed to get tokens: {response.text}")

def refresh_access_token(refresh_token):
    """Refresh the access token."""
    import requests

    response = requests.post(
        "https://api.box.com/oauth2/token",
        data={
            "grant_type": "refresh_token",
            "refresh_token": refresh_token,
            "client_id": BOX_CLIENT_ID,
            "client_secret": BOX_CLIENT_SECRET,
        }
    )

    if response.status_code == 200:
        data = response.json()
        return data["access_token"], data["refresh_token"]
    else:
        return None, None

def upload_to_box(access_token, file_data, filename, month_number, year):
    """Upload file to Box, creating year and month folders if needed."""
    import requests

    headers = {"Authorization": f"Bearer {access_token}"}

    # Step 1: Find or create year folder
    response = requests.get(
        f"https://api.box.com/2.0/folders/{BOX_ROOT_FOLDER_ID}/items",
        headers=headers
    )

    if response.status_code != 200:
        raise Exception(f"Cannot access root folder (status {response.status_code}): {response.text}")

    year_folder_id = None
    items = response.json().get("entries", [])
    for item in items:
        if item["type"] == "folder" and item["name"] == year:
            year_folder_id = item["id"]
            break

    # Create year folder if it doesn't exist
    if not year_folder_id:
        response = requests.post(
            "https://api.box.com/2.0/folders",
            headers={**headers, "Content-Type": "application/json"},
            json={"name": year, "parent": {"id": BOX_ROOT_FOLDER_ID}}
        )
        if response.status_code == 201:
            year_folder_id = response.json()["id"]
        elif response.status_code == 409:
            # Folder already exists - refetch
            response = requests.get(f"https://api.box.com/2.0/folders/{BOX_ROOT_FOLDER_ID}/items", headers=headers)
            for item in response.json().get("entries", []):
                if item["type"] == "folder" and item["name"] == year:
                    year_folder_id = item["id"]
                    break
        else:
            raise Exception(f"Failed to create year folder (status {response.status_code}): {response.text}")

    # Step 2: Find or create month folder inside year folder
    month_name = MONTH_NAMES.get(month_number, "Unknown")
    folder_name = f"{month_number} {month_name}"

    response = requests.get(
        f"https://api.box.com/2.0/folders/{year_folder_id}/items",
        headers=headers
    )

    month_folder_id = None
    if response.status_code == 200:
        items = response.json().get("entries", [])
        for item in items:
            if item["type"] == "folder" and item["name"] == folder_name:
                month_folder_id = item["id"]
                break

    # Create month folder if it doesn't exist
    if not month_folder_id:
        response = requests.post(
            "https://api.box.com/2.0/folders",
            headers={**headers, "Content-Type": "application/json"},
            json={"name": folder_name, "parent": {"id": year_folder_id}}
        )
        if response.status_code == 201:
            month_folder_id = response.json()["id"]
        elif response.status_code == 409:
            # Folder already exists - find it
            for item in items:
                if item["type"] == "folder" and item["name"] == folder_name:
                    month_folder_id = item["id"]
                    break
        else:
            raise Exception(f"Failed to create month folder (status {response.status_code}): {response.text}")

    # Upload file
    response = requests.post(
        "https://upload.box.com/api/2.0/files/content",
        headers=headers,
        data={"attributes": json.dumps({"name": filename, "parent": {"id": month_folder_id}})},
        files={"file": (filename, file_data)}
    )

    if response.status_code == 201:
        return response.json(), folder_name, month_folder_id
    elif response.status_code == 409:
        # File already exists - upload with a new name (add timestamp)
        from datetime import datetime
        name_parts = filename.rsplit('.', 1)
        timestamp = datetime.now().strftime("%H%M%S")
        if len(name_parts) == 2:
            new_filename = f"{name_parts[0]}_{timestamp}.{name_parts[1]}"
        else:
            new_filename = f"{filename}_{timestamp}"

        # Try uploading with new name
        retry_response = requests.post(
            "https://upload.box.com/api/2.0/files/content",
            headers=headers,
            data={"attributes": json.dumps({"name": new_filename, "parent": {"id": month_folder_id}})},
            files={"file": (new_filename, file_data)}
        )
        if retry_response.status_code == 201:
            return retry_response.json(), folder_name, month_folder_id
        else:
            return {"status": "uploaded as " + new_filename}, folder_name, month_folder_id
    else:
        raise Exception(f"Failed to upload (status {response.status_code}): {response.text}")

# Check if we got a code from Box redirect
query_params = st.query_params
if "code" in query_params:
    auth_code = query_params["code"]
    try:
        access_token, refresh_token = exchange_code_for_tokens(auth_code)
        # Get Box user info
        user_info = get_box_user_info(access_token)
        if user_info:
            # Save user info to session state
            st.session_state["box_user_id"] = user_info["id"]
            st.session_state["box_user_name"] = user_info["name"]
            st.session_state["box_user_email"] = user_info["email"]
            # Save tokens with user info
            save_tokens(access_token, refresh_token, user_info["id"], user_info["name"], user_info["email"])
        st.query_params.clear()
        st.rerun()
    except Exception as e:
        st.error(f"Failed to connect: {str(e)}")
        st.query_params.clear()

# Sidebar for Box auth
with st.sidebar:
    st.header("Box Connection")

    # Check for Box connection
    box_connected = False
    tokens = None
    box_user_name = st.session_state.get("box_user_name")

    # If we have a box_user_id in session, try to load/refresh tokens
    if st.session_state.get("box_user_id"):
        tokens = load_tokens()
        if tokens and tokens.get("refresh_token"):
            # Try to refresh token (Box tokens expire after 60 min)
            new_access, new_refresh = refresh_access_token(tokens["refresh_token"])
            if new_access:
                save_tokens(new_access, new_refresh, st.session_state["box_user_id"],
                           st.session_state.get("box_user_name"), st.session_state.get("box_user_email"))
                tokens = {"access_token": new_access, "refresh_token": new_refresh}
                box_connected = True
            else:
                # Refresh failed - token expired, need to reconnect
                box_connected = False
        elif tokens and tokens.get("access_token"):
            box_connected = True

    if box_connected and box_user_name:
        st.success(f"Connected as {box_user_name}")
        if st.button("Disconnect"):
            delete_tokens()
            st.session_state.pop("box_user_id", None)
            st.session_state.pop("box_user_name", None)
            st.session_state.pop("box_user_email", None)
            st.rerun()
    else:
        st.warning("Not connected to Box")
        st.caption("Click below to log in with your Box account")
        auth_url = f"https://account.box.com/api/oauth2/authorize?client_id={BOX_CLIENT_ID}&redirect_uri={BOX_REDIRECT_URI}&response_type=code"
        st.link_button("Log in with Box", auth_url)

# Display logo centered
logo_path = Path(__file__).parent / "logo.png"
if logo_path.exists():
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        st.image(str(logo_path), width=250)

st.markdown("<h1 style='text-align: center;'>Financial Package Aggregator</h1>", unsafe_allow_html=True)

# Custom CSS for better styling
st.markdown("""
<style>
    .property-card {
        background: linear-gradient(135deg, #f8f9fa 0%, #e9ecef 100%);
        border-radius: 12px;
        padding: 1.2rem;
        margin: 0.8rem 0;
        border-left: 4px solid #4CAF50;
        box-shadow: 0 2px 4px rgba(0,0,0,0.1);
    }
    .property-card.incomplete {
        border-left-color: #1976d2;
    }
    .property-title {
        font-size: 1.1rem;
        font-weight: 600;
        color: #1a1a2e;
        margin-bottom: 0.5rem;
    }
    .report-grid {
        display: grid;
        grid-template-columns: repeat(auto-fill, minmax(200px, 1fr));
        gap: 0.5rem;
        margin-top: 0.8rem;
    }
    .report-item {
        background: white;
        padding: 0.5rem 0.8rem;
        border-radius: 6px;
        font-size: 0.85rem;
        display: flex;
        align-items: center;
        gap: 0.5rem;
    }
    .report-item.found { border-left: 3px solid #4CAF50; }
    .report-item.missing { border-left: 3px solid #f44336; background: #fff5f5; }
    .status-badge {
        display: inline-block;
        padding: 0.25rem 0.6rem;
        border-radius: 20px;
        font-size: 0.75rem;
        font-weight: 500;
    }
    .badge-complete { background: #e8f5e9; color: #2e7d32; }
    .badge-incomplete { background: #e3f2fd; color: #1565c0; }
    .result-card {
        background: white;
        border-radius: 10px;
        padding: 1rem;
        margin: 0.5rem 0;
        border: 1px solid #e0e0e0;
        display: flex;
        justify-content: space-between;
        align-items: center;
    }
    .result-card.success { border-left: 4px solid #4CAF50; }
    .result-card.error { border-left: 4px solid #f44336; }
</style>
""", unsafe_allow_html=True)

st.markdown("---")

# Keywords to identify each report type (in merge order)
report_patterns = [
    (1, "Balance Sheet", ["balance_sheet", "balance sheet", "balancesheet", "period_change", "periodchange"]),
    (2, "T-12 Statement", ["12_month", "12 month", "12month", "t-12", "t12", "trailing"]),
    (3, "YTD Statement", ["ytd", "year_to_date", "year-to-date", "yeartodate"]),
    (4, "Budget Comparison", ["budget", "comparison"]),
    (5, "Rent Roll", ["renrollwithleasecharges", "rent_roll", "rentroll", "lease_charges"]),
    (6, "Aged Receivables", ["aging_summary", "agingsummary"]),
    (7, "Payables Aging", ["payablesagingreport", "payablesaging"]),
    (8, "General Ledger", ["general_ledger", "generalledger", "gl_detail", "gldetail"]),
]

# Properties that require Excel file with T-12, YTD, and General Ledger
PROPERTIES_WITH_EXCEL = {"marshp", "emersn", "capella2", "55pharr"}

# Map property codes to full property names
PROPERTY_NAMES = {
    "marshp": "Marsh Point",
    "emersn": "Emerson",
    "capella2": "Capella",
    "55pharr": "55 Pharr",
}

def identify_report(filename):
    """Identify report type based on filename."""
    filename_lower = filename.lower()
    for order, name, keywords in report_patterns:
        for keyword in keywords:
            if keyword in filename_lower:
                return (order, name)
    return (99, "Unknown")

# File uploader for PDFs and Excel files
uploaded_files = st.file_uploader(
    "Upload PDF reports and Excel files",
    type=["pdf", "xlsx", "xls"],
    accept_multiple_files=True
)

st.markdown("---")

# Initialize properties dict (used across sections)
properties = defaultdict(list)
property_codes = {}  # Map property name to property code
excel_files = defaultdict(dict)  # Excel files grouped by property code {code: {"T-12": bytes, "YTD": bytes, "GL": bytes}}
unidentified_property = []
unidentified_excel = []

# Month and Year selection
col1, col2 = st.columns(2)
with col1:
    month_number = st.selectbox("Month", options=list(MONTH_NAMES.keys()), format_func=lambda x: f"{x} - {MONTH_NAMES[x]}")
with col2:
    year = st.text_input("Year", value="2026")

# Process uploaded files
if uploaded_files:
    with st.spinner("Analyzing uploaded files..."):
        # Group files by property
        properties = defaultdict(list)
        property_codes = {}  # Map property name to property code
        excel_files = defaultdict(dict)  # Excel files grouped by property code
        unidentified_property = []
        unidentified_excel = []

        for file in uploaded_files:
            file_bytes = file.read()
            file.seek(0)  # Reset for later use

            # Check if it's an Excel file
            if file.name.lower().endswith(('.xlsx', '.xls')):
                filename_lower = file.name.lower()

                # Extract property code from filename
                # First, search for known property codes anywhere in the filename
                prop_code = None
                for code in PROPERTIES_WITH_EXCEL:
                    if f"_{code}" in filename_lower or f"_{code}." in filename_lower or f"_{code} " in filename_lower:
                        prop_code = code
                        break

                # Fallback to regex pattern if no known code found
                if not prop_code:
                    # Handle double extensions like .xls.xlsx
                    clean_filename = re.sub(r'\.(xls\.xlsx|xlsx|xls)$', '', filename_lower)
                    # Remove trailing report type indicators (t12, ytd, etc.) before extracting property code
                    clean_filename = re.sub(r'[\s_]*(t12|ytd|t-12)[\s]*(\(\d+\))?$', '', clean_filename, flags=re.IGNORECASE)
                    code_match = re.search(r'_([a-z0-9]+)(?:\s*\(\d+\))?$', clean_filename)
                    prop_code = code_match.group(1) if code_match else None

                # Identify report type using same patterns as PDFs
                order, report_type = identify_report(file.name)

                # Check for YTD or T12 at end of filename to override detection
                # This handles cases like "12_Month_Statement_marshp (1) YTD.xlsx"
                name_without_ext = re.sub(r'\.(xlsx|xls)$', '', filename_lower)
                if name_without_ext.endswith('ytd') or ' ytd' in name_without_ext:
                    order = 3
                    report_type = "YTD Statement"
                elif name_without_ext.endswith('t12') or ' t12' in name_without_ext:
                    order = 2
                    report_type = "T-12 Statement"

                # For special properties, track T-12, YTD, GL Excel files for merged Excel
                if prop_code in PROPERTIES_WITH_EXCEL:
                    if report_type == "T-12 Statement":
                        excel_files[prop_code]["T-12"] = {"bytes": file_bytes, "filename": file.name}
                    elif report_type == "YTD Statement":
                        excel_files[prop_code]["YTD"] = {"bytes": file_bytes, "filename": file.name}
                    elif report_type == "General Ledger":
                        excel_files[prop_code]["GL"] = {"bytes": file_bytes, "filename": file.name}

                # Convert Excel to PDF for merging (skip General Ledger for PDF merge)
                if report_type != "General Ledger" and order != 99:
                    try:
                        pdf_bytes = excel_to_pdf(file_bytes)
                        prop_name = PROPERTY_NAMES.get(prop_code, prop_code.title() if prop_code else None)

                        if prop_name:
                            # Check if this property code already exists under a different name
                            existing_name = None
                            for name, code in property_codes.items():
                                if code == prop_code:
                                    existing_name = name
                                    break
                            if existing_name:
                                prop_name = existing_name

                            properties[prop_name].append({
                                "file": file,
                                "bytes": pdf_bytes,
                                "order": order,
                                "report_type": report_type,
                                "from_excel": True
                            })
                            if prop_code:
                                property_codes[prop_name] = prop_code
                        else:
                            unidentified_property.append({
                                "file": file,
                                "bytes": pdf_bytes,
                                "order": order,
                                "report_type": report_type
                            })
                    except Exception as e:
                        st.warning(f"Could not convert {file.name} to PDF: {e}")
                elif report_type == "General Ledger":
                    pass  # GL only goes to Excel, not PDF
                else:
                    unidentified_excel.append({
                        "filename": file.name,
                        "property_code": prop_code,
                        "report_type": report_type
                    })
                continue

            # For PDFs: Extract property name and code
            prop_name, prop_code = extract_property_info(file_bytes)

            # Fallback: extract property code from filename (e.g., "Report_marshp.pdf")
            if not prop_code:
                filename_lower = file.name.lower()
                # First, search for known property codes anywhere in the filename
                for code in PROPERTIES_WITH_EXCEL:
                    if f"_{code}" in filename_lower or f"_{code}." in filename_lower or f"_{code} " in filename_lower:
                        prop_code = code
                        break

                # Fallback to regex pattern if no known code found
                if not prop_code:
                    # Remove trailing report type indicators before extracting property code
                    clean_filename = re.sub(r'\.pdf$', '', filename_lower)
                    clean_filename = re.sub(r'[\s_]*(t12|ytd|t-12)[\s]*(\(\d+\))?$', '', clean_filename, flags=re.IGNORECASE)
                    code_match = re.search(r'_([a-z0-9]+)(?:\s*\(\d+\))?$', clean_filename)
                    if code_match:
                        prop_code = code_match.group(1)

            # Use property code as property name if we couldn't extract from PDF
            if not prop_name and prop_code:
                # Map common codes to property names
                prop_name = PROPERTY_NAMES.get(prop_code, prop_code.title())

            # Identify report type
            order, report_type = identify_report(file.name)

            # If it's a 12-month statement, check PDF content to determine T-12 vs YTD
            if report_type == "T-12 Statement":
                statement_type = is_t12_or_ytd(file_bytes)
                if statement_type == "YTD":
                    order = 3
                    report_type = "YTD Statement"

            if prop_name:
                # Normalize property name - use property code to group if we have it
                if prop_code:
                    # Check if this property code already exists under a different name
                    existing_name = None
                    for name, code in property_codes.items():
                        if code == prop_code:
                            existing_name = name
                            break
                    if existing_name:
                        prop_name = existing_name

                properties[prop_name].append({
                    "file": file,
                    "bytes": file_bytes,
                    "order": order,
                    "report_type": report_type
                })
                if prop_code:
                    property_codes[prop_name] = prop_code
            else:
                unidentified_property.append({
                    "file": file,
                    "bytes": file_bytes,
                    "order": order,
                    "report_type": report_type
                })

    # Display grouped results

    if unidentified_property:
        with st.expander(f"⚠️ {len(unidentified_property)} unmatched file(s)", expanded=False):
            for item in unidentified_property:
                st.write(f"• {item['file'].name}")

    # Expected report types
    base_expected_reports = {"Balance Sheet", "T-12 Statement", "YTD Statement", "Budget Comparison", "Rent Roll", "Aged Receivables", "Payables Aging"}

    # Display each property's reports
    for prop_name, files in sorted(properties.items()):
        # Sort files by report order
        files.sort(key=lambda x: x["order"])

        # Check if this property needs Excel file (and therefore General Ledger)
        prop_code = property_codes.get(prop_name, "")
        needs_excel = prop_code in PROPERTIES_WITH_EXCEL

        if needs_excel:
            expected_reports = base_expected_reports | {"General Ledger"}
            expected_count = 8
        else:
            expected_reports = base_expected_reports
            expected_count = 7

        # Check for missing reports
        found_reports = {f["report_type"] for f in files if f["report_type"] != "Unknown"}

        # For special properties, Excel uploads also count towards the reports
        if needs_excel:
            excel_data = excel_files.get(prop_code, {})
            if "T-12" in excel_data:
                found_reports.add("T-12 Statement")
            if "YTD" in excel_data:
                found_reports.add("YTD Statement")
            if "GL" in excel_data:
                found_reports.add("General Ledger")

        missing_reports = expected_reports - found_reports
        is_complete = len(missing_reports) == 0

        # Status badge
        if is_complete:
            badge = '<span class="status-badge badge-complete">Complete</span>'
            card_class = "property-card"
        else:
            badge = f'<span class="status-badge badge-incomplete">{len(missing_reports)} Missing</span>'
            card_class = "property-card incomplete"

        # Build directory-style tree
        tree_lines = []
        all_reports = list(expected_reports)
        all_reports.sort()

        for i, report in enumerate(all_reports):
            is_last = (i == len(all_reports) - 1)
            prefix = "└──" if is_last else "├──"

            if report in found_reports:
                tree_lines.append(f'<div style="color:#2e7d32;">{prefix} {report}</div>')
            else:
                tree_lines.append(f'<div style="color:#bbb;">{prefix} {report}</div>')

        tree_html = "\n".join(tree_lines)

        st.markdown(f"""
        <div class="{card_class}">
            <div class="property-title">{prop_name} {badge}</div>
            <div style="font-family:monospace;font-size:0.85rem;margin-top:0.5rem;line-height:1.6;">
                {tree_html}
            </div>
        </div>
        """, unsafe_allow_html=True)

    # Display unidentified Excel files
    if unidentified_excel:
        with st.expander(f"⚠️ {len(unidentified_excel)} unidentified Excel file(s)", expanded=False):
            for item in unidentified_excel:
                st.write(f"• {item['filename']}")

st.markdown("---")

# Merge and Upload button
if st.button("Merge & Upload to Box", type="primary", use_container_width=True):
    if not uploaded_files:
        st.error("Please upload files first.")
    elif not box_connected:
        st.error("Please connect to Box first (see sidebar).")
    elif not properties:
        st.error("No properties detected in uploaded files.")
    else:
        try:
            results = []
            excel_results = []

            progress_bar = st.progress(0)
            status_text = st.empty()

            # Clear previous results
            st.session_state["upload_results"] = None
            st.session_state["excel_results"] = None

            total_properties = len(properties)

            for idx, (prop_name, files) in enumerate(sorted(properties.items())):
                status_text.text(f"Processing {prop_name}...")

                # Sort files by report order
                files.sort(key=lambda x: x["order"])

                # Check if this is a special property that needs Excel
                prop_code = property_codes.get(prop_name, "")
                needs_excel = prop_code in PROPERTIES_WITH_EXCEL

                # Separate General Ledger from other reports
                pdf_files = []
                t12_bytes = None
                ytd_bytes = None
                gl_bytes = None

                for item in files:
                    if item["report_type"] == "General Ledger":
                        gl_bytes = item["bytes"]
                    else:
                        pdf_files.append(item)
                        if item["report_type"] == "T-12 Statement":
                            t12_bytes = item["bytes"]
                        elif item["report_type"] == "YTD Statement":
                            ytd_bytes = item["bytes"]

                # Merge PDFs (exclude General Ledger)
                merger = PdfMerger()
                for item in pdf_files:
                    merger.append(io.BytesIO(item["bytes"]))

                # Create PDF output
                output = io.BytesIO()
                merger.write(output)
                merger.close()
                output.seek(0)

                # Generate filename: "{Property Name} Financials {MM} {YYYY}.pdf"
                pdf_filename = f"{prop_name} Financials {month_number} {year}.pdf"

                # Upload PDF to Box
                try:
                    tokens = load_tokens()
                    if not tokens:
                        raise Exception("No Box tokens found - please reconnect to Box")
                    st.info(f"Uploading {pdf_filename}...")
                    uploaded_result, folder_name, month_folder_id = upload_to_box(
                        tokens["access_token"],
                        output.getvalue(),
                        pdf_filename,
                        month_number,
                        year
                    )
                    st.info(f"Upload result: {uploaded_result.get('status', 'success')}")
                    # Extract file ID from upload response
                    file_id = None
                    if isinstance(uploaded_result, dict) and "entries" in uploaded_result:
                        file_id = uploaded_result["entries"][0]["id"]
                    results.append({
                        "property": prop_name,
                        "filename": pdf_filename,
                        "data": output.getvalue(),
                        "folder": folder_name,
                        "folder_id": month_folder_id,
                        "file_id": file_id,
                        "status": "success"
                    })
                except Exception as e:
                    results.append({
                        "property": prop_name,
                        "filename": pdf_filename,
                        "data": output.getvalue(),
                        "status": "error",
                        "error": str(e)
                    })

                # Create merged Excel for special properties (from uploaded Excel files)
                if needs_excel and prop_code in excel_files:
                    excel_data_for_prop = excel_files[prop_code]
                    t12_excel = excel_data_for_prop.get("T-12", {}).get("bytes")
                    ytd_excel = excel_data_for_prop.get("YTD", {}).get("bytes")
                    gl_excel = excel_data_for_prop.get("GL", {}).get("bytes")

                    if t12_excel and ytd_excel and gl_excel:
                        status_text.text(f"Creating Excel for {prop_name}...")
                        excel_data = merge_excel_files(t12_excel, ytd_excel, gl_excel)
                        excel_filename = f"{prop_name} Financials {month_number} {year}.xlsx"

                        # Upload Excel to Box
                        try:
                            tokens = load_tokens()
                            uploaded_result, folder_name, month_folder_id = upload_to_box(
                                tokens["access_token"],
                                excel_data,
                                excel_filename,
                                month_number,
                                year
                            )
                            # Extract file ID from upload response
                            file_id = None
                            if isinstance(uploaded_result, dict) and "entries" in uploaded_result:
                                file_id = uploaded_result["entries"][0]["id"]
                            excel_results.append({
                                "property": prop_name,
                                "filename": excel_filename,
                                "data": excel_data,
                                "folder": folder_name,
                                "folder_id": month_folder_id,
                                "file_id": file_id,
                                "status": "success"
                            })
                        except Exception as e:
                            excel_results.append({
                                "property": prop_name,
                                "filename": excel_filename,
                                "data": excel_data,
                                "status": "error",
                                "error": str(e)
                            })
                    else:
                        missing = []
                        if not t12_excel: missing.append("T-12")
                        if not ytd_excel: missing.append("YTD")
                        if not gl_excel: missing.append("General Ledger")
                        st.warning(f"Could not create Excel for {prop_name}: missing Excel files for {', '.join(missing)}")
                elif needs_excel:
                    st.warning(f"No Excel files uploaded for {prop_name}")

                progress_bar.progress((idx + 1) / total_properties)

            status_text.empty()
            progress_bar.empty()

            # Save results to session state so they persist
            st.session_state["upload_results"] = results
            st.session_state["excel_results"] = excel_results
            st.session_state["folder_id"] = results[0].get("folder_id") if results else None
            st.session_state["folder_name"] = results[0].get("folder") if results else None
            st.session_state["upload_year"] = year

        except Exception as e:
            st.error(f"Error: {str(e)}")
            import traceback
            st.code(traceback.format_exc())

# Display results from session state (persists across reruns)
if st.session_state.get("upload_results"):
    results = st.session_state["upload_results"]
    excel_results = st.session_state.get("excel_results", [])
    folder_id = st.session_state.get("folder_id")
    folder_name = st.session_state.get("folder_name")
    upload_year = st.session_state.get("upload_year")

    st.success(f"Uploaded to Box: {upload_year} / {folder_name}")
    if folder_id:
        st.markdown(f"[Open folder in Box](https://app.box.com/folder/{folder_id})")

    st.markdown("---")

    # Show PDF downloads
    st.markdown("**PDF Packages**")
    for i, result in enumerate(results):
        col1, col2 = st.columns([4, 1])
        with col1:
            if result.get("status") == "success":
                st.markdown(f"{result['property']}")
            else:
                st.error(f"{result['property']}: {result.get('error', 'Unknown error')}")
        with col2:
            st.download_button(
                label="⬇",
                data=result["data"],
                file_name=result["filename"],
                mime="application/pdf",
                key=f"download_pdf_{i}"
            )

    # Show Excel downloads
    if excel_results:
        st.markdown("")
        st.markdown("**Excel Packages**")
        for i, result in enumerate(excel_results):
            col1, col2 = st.columns([4, 1])
            with col1:
                if result.get("status") == "success":
                    st.markdown(f"{result['property']}")
                else:
                    st.error(f"{result['property']}: {result.get('error', 'Unknown error')}")
            with col2:
                st.download_button(
                    label="⬇",
                    data=result["data"],
                    file_name=result["filename"],
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"download_excel_{i}"
                )
